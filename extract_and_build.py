"""
extract_and_build.py

Pipeline completo: conecta na montagem aberta no SolidWorks, extrai cada
componente, e gera as abas correspondentes no Cost Report/BOM.

Como rodar (no Windows, com a montagem aberta no SolidWorks):
    pip install pywin32
    python extract_and_build.py --template caminho\\para\\template.xlsx --out caminho\\para\\saida.xlsx

NAO TESTADO em maquina real -- rodem primeiro sw_extract/test_connection.py
numa peca simples antes de tentar isto numa montagem inteira. Esperem
precisar ajustar nomes de metodo/assinaturas dependendo da versao do
SolidWorks instalada.
"""

import argparse
import sys

import openpyxl

from bom_builder import create_part_sheet, add_part_to_bom
from sw_extract.connector import connect_to_solidworks, get_active_document, SW_DOC_ASSEMBLY
from sw_extract.extractor import traverse_assembly
from sw_extract.material_matcher import load_catalog_material_names
from sw_extract.translator import build_part_dict


def next_sheet_code(wb, system_prefix):
    """
    Gera o proximo codigo de aba disponivel para um prefixo de sistema
    (ex: 'BR' -> 'BR30' se BR01..BR29 ja existirem). Ajustem os prefixos
    abaixo para bater com a convencao real do time.
    """
    existing_nums = []
    for name in wb.sheetnames:
        if name.startswith(system_prefix) and name[len(system_prefix):].isdigit():
            existing_nums.append(int(name[len(system_prefix):]))
    next_num = (max(existing_nums) + 1) if existing_nums else 1
    return f"{system_prefix}{next_num:02d}"


SYSTEM_PREFIX_MAP = {
    "Brake System": "BR",
    "Drivetrain": "DT",
    "Frame & Body": "FR",
    "Electrical": "EL",
    "Miscellaneous, Fit & Finish": "MS",
    "Steering System": "ST",
    "Suspension": "SU",
    "Wheels & Tires": "WT",
}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, help="Cost Report .xlsx base")
    parser.add_argument("--out", required=True, help="Arquivo de saida .xlsx")
    parser.add_argument("--university", default="")
    args = parser.parse_args()

    print("Conectando ao SolidWorks...")
    sw_app = connect_to_solidworks()
    model = get_active_document(sw_app)
    if model.GetType != SW_DOC_ASSEMBLY:  # sem parenteses -- ver aviso em sw_extract/connector.py
        print("ERRO: o documento ativo nao e uma montagem (assembly). "
              "Abram a montagem que querem extrair e deixem em foco.")
        sys.exit(1)

    print("Percorrendo a arvore de componentes...")
    components = traverse_assembly(model)
    print(f"  {len(components)} componentes unicos encontrados.")

    print("Carregando template do Cost Report...")
    wb = openpyxl.load_workbook(args.template, data_only=False)
    catalog_material_names = load_catalog_material_names(wb)

    all_warnings = []
    generated = []
    failed = []

    for entry in components:
        part, warnings = build_part_dict(entry, catalog_material_names,
                                          university=args.university)
        all_warnings.extend(warnings)
        if part is None:
            continue
        if not part["system"]:
            all_warnings.append(
                f"Peca '{part['part_name']}' pulada -- sem FSAE_System "
                f"definido, nao sei em qual sistema/aba registrar."
            )
            continue

        prefix = SYSTEM_PREFIX_MAP.get(part["system"])
        if prefix is None:
            all_warnings.append(
                f"Peca '{part['part_name']}': sistema '{part['system']}' "
                f"nao reconhecido (confira SYSTEM_PREFIX_MAP) -- pulada."
            )
            continue

        sheet_name = next_sheet_code(wb, prefix)
        if not part["pn_base"]:
            part["pn_base"] = f"FSAEB-26-24-{sheet_name}"

        # Uma peca com problema (ex.: nome de sistema com erro de digitacao)
        # nao pode derrubar o lote inteiro -- registra a falha e continua,
        # pra nao perder as pecas ja processadas com sucesso antes dela.
        try:
            subtotals = create_part_sheet(wb, template_sheet_name="BR01",
                                           new_sheet_name=sheet_name, part=part)
            all_warnings.extend(subtotals.get("warnings", []))
            add_part_to_bom(wb, part, subtotals)
        except Exception as exc:
            failed.append(
                f"Peca '{part['part_name']}' (aba {sheet_name}) falhou ao "
                f"gerar/registrar na BOM: {exc}"
            )
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            print(f"  ! FALHOU {sheet_name}: {part['part_name']} -- {exc}")
            continue

        generated.append(sheet_name)
        print(f"  + {sheet_name}: {part['part_name']}")

    wb.save(args.out)
    print(f"\nSalvo em: {args.out}")
    print(f"{len(generated)} pecas geradas: {generated}")

    if all_warnings:
        print(f"\n{len(all_warnings)} avisos -- revisar antes de submeter:")
        for w in all_warnings:
            print("  -", w)

    if failed:
        print(f"\n{len(failed)} pecas FALHARAM e nao foram incluidas:")
        for f in failed:
            print("  -", f)


if __name__ == "__main__":
    main()
