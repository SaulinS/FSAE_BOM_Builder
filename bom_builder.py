"""
fsae_bom / bom_builder.py

Prototipo: gera uma nova aba de peca no Cost Report/BOM da Formula SAE Brasil
(no formato exato usado pela equipe, validado a partir do arquivo real
24_FSAEB2026_CReBOM_SENAICIMATEC_TEC_Racing.xlsx) e registra essa peca
corretamente na aba BOM mestra e na aba Cost Summary.

ESCOPO DESTE PROTOTIPO (combinado com o time):
- Preenche automaticamente: cabecalho da peca, bloco de Materiais e bloco de
  Processos (com as formulas de VLOOKUP contra os catalogos Materials/
  Processes/Process Multipliers, exatamente como o arquivo original faz).
- Os blocos de Fixadores (Fasteners) e Ferramental (Tooling) sao criados com
  cabecalho e linhas em branco para preenchimento manual (como e feito hoje),
  ja que a extracao automatica desses itens fica para uma fase futura.
- O custo unitario de Material (coluna UnitCost) e preenchido automaticamente
  sempre que o nome do material e encontrado no catalogo "Materials": para
  materiais de custo fixo (ex.: itens comprados) gera um VLOOKUP direto na
  coluna Cost; para materiais com formula (ligada a Size1/Size2/Area/Length/
  Density) reconstroi essa formula substituindo os tokens [C1]/[C2]/[Size1]/
  [Size2]/[Area]/[Length]/[Density] por VLOOKUPs/referencias reais (ver
  _material_unitcost_formula). So fica em branco quando o nome do material
  nao e encontrado no catalogo, ou e encontrado mas nao tem nem formula nem
  custo fixo definidos -- nesses casos entra como aviso na lista de warnings
  retornada por create_part_sheet.
- Quando a formula do catalogo exige uma dimensao (Size1/Size2/Area/Length/
  Density), a formula gerada sai protegida por uma guarda que devolve NA()
  se a celula correspondente estiver vazia, e a falta tambem entra como
  aviso na geracao. Sem isso, uma dimensao nao preenchida faria o Excel
  multiplicar por celula vazia (= 0) e o material entraria custando 0,00
  sem erro nenhum -- ver _material_unitcost_formula para o porque e para os
  numeros medidos no catalogo real.

LIMITACAO CONHECIDA (documentada, nao escondida):
- A insercao de uma peca nova desloca linhas da aba BOM. O script recalcula
  automaticamente os totais por sistema (SUMPRODUCT) e o Total do Veiculo,
  alem de corrigir as referencias na aba "Cost Summary" -- mas caso a estrutura
  oficial do template mude de um ano para o outro (nomes de aba, ordem de
  colunas), este script precisa ser revisado.
"""

import copy
import re
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# Varias linhas da aba BOM usam formulas que "auto-referenciam" o proprio
# numero de linha em texto -- ex.: =INDIRECT(C123&"!B4") (colunas D/E/F/G/I),
# =SUM(J123:M123) (coluna H, Unit Cost) e =H123*I123 (coluna N, Total Cost).
# Isso e fragil: ao inserir uma linha nova em QUALQUER lugar da tabela, todas
# as linhas abaixo sao deslocadas fisicamente, mas o numero "123" dentro do
# texto da formula NAO acompanha o deslocamento (limitacao do openpyxl/Excel
# ao inserir linhas). O resultado e que a formula passa a apontar para a
# linha errada -- silenciosamente. Isso ja era uma fragilidade existente no
# arquivo original (nao introduzida por este script), mas nossa insercao de
# linhas expõe o problema, entao corrigimos automaticamente apos cada
# insercao -- nas tres formas conhecidas de auto-referencia acima.
_INDIRECT_SELF_REF_PATTERN = re.compile(r'C(\d+)&')
_UNITCOST_SELF_REF_PATTERN = re.compile(r'=SUM\(J(\d+):M(\d+)\)')
_TOTALCOST_SELF_REF_PATTERN = re.compile(r'=H(\d+)\*I(\d+)')


# ---------------------------------------------------------------------------
# Utilidades de estilo
# ---------------------------------------------------------------------------

def _clone_cell_style(dst_cell, src_cell):
    """Copia fonte, preenchimento, borda, alinhamento e formato numerico."""
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def _copy_row_style(ws_src, src_row, ws_dst, dst_row, max_col=14):
    for col in range(1, max_col + 1):
        _clone_cell_style(ws_dst.cell(row=dst_row, column=col),
                           ws_src.cell(row=src_row, column=col))


# ---------------------------------------------------------------------------
# Catalogo de Materiais: a aba "Materials" documenta, por item, a formula de
# custo (colunas [C1], [C2], Formula, Cost). Isso permite montar a formula de
# UnitCost automaticamente na aba da peca, em vez de deixar em branco.
# ---------------------------------------------------------------------------

def _load_materials_catalog(wb):
    ws = wb["Materials"]
    catalog = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=14):
        name = row[2].value  # coluna C
        if not name:
            continue
        catalog[name] = {"formula": row[11].value, "cost": row[12].value}
    return catalog


def _load_catalog_names(wb, sheet_name, name_col_idx):
    ws = wb[sheet_name]
    names = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        v = row[name_col_idx].value
        if v:
            names.add(v)
    return names


# Tokens da coluna 'Formula' do catalogo que dependem de um valor
# preenchido NA LINHA DA PECA (nao de um coeficiente do catalogo), com a
# coluna da aba de peca onde esse valor mora e o campo correspondente do
# dicionario `part` que alimenta create_part_sheet.
_DIMENSION_TOKENS = {
    "[Size1]":   ("E", "size1"),
    "[Size2]":   ("G", "size2"),
    "[Area]":    ("J", "area"),
    "[Length]":  ("K", "length"),
    "[Density]": ("L", "density"),
}


def _material_unitcost_formula(materials_catalog, material_name, row):
    """
    Monta a formula de UnitCost de uma linha de material substituindo os
    tokens [C1]/[C2]/[Size1]/[Size2]/[Area]/[Length]/[Density] da coluna
    'Formula' do catalogo por referencias reais (VLOOKUP para os
    coeficientes, celulas da propria linha para as dimensoes).

    Retorna (formula, tokens_exigidos), onde `tokens_exigidos` sao os
    tokens de dimensao que a formula usa -- quem chama precisa disso para
    conferir se a peca realmente forneceu esses campos. A formula e None se
    o material nao for encontrado no catalogo ou nao tiver formula/custo.

    PROTECAO CONTRA CUSTO ZERO SILENCIOSO: uma dimensao que a formula exige
    mas ninguem preencheu deixa a celula vazia, e o Excel trata celula vazia
    como 0 numa multiplicacao -- o custo do material vira 0,00 sem erro
    nenhum, e o subtotal so fica "baixo" em vez de obviamente quebrado.
    Medido no catalogo real da FSAE Brasil 2026, 119 dos 1091 materiais
    (10,9%) caem nesse caso pela extracao automatica, quase todos por causa
    de [Size2]. Por isso a formula sai embrulhada numa guarda que devolve
    NA() quando qualquer celula exigida estiver vazia: #N/A propaga pelo
    SUM, entao o subtotal e o total do veiculo quebram de forma visivel em
    vez de mentir para baixo. A guarda tambem protege contra alguem apagar
    o valor depois, no Excel -- o que a validacao em tempo de geracao,
    sozinha, nao pegaria.
    """
    info = materials_catalog.get(material_name)
    if info is None:
        return None, []

    cost = info["cost"]
    if isinstance(cost, (int, float)):
        # Custo fixo (ex.: item comprado, "unit"): so precisa da VLOOKUP,
        # nao depende de nenhuma dimensao da linha.
        return f"=VLOOKUP(B{row},Materials!C:M,11,FALSE)", []

    formula_text = info["formula"]
    if not formula_text or not isinstance(formula_text, str):
        return None, []

    expr = formula_text.lstrip("=")
    c1_ref = f"VLOOKUP(B{row},Materials!C:J,8,FALSE)"
    c2_ref = f"VLOOKUP(B{row},Materials!C:K,9,FALSE)"
    expr = expr.replace("[C1]", f"({c1_ref})")
    expr = expr.replace("[C2]", f"({c2_ref})")

    required = []
    for token, (col, _field) in _DIMENSION_TOKENS.items():
        if token in expr:
            required.append(token)
            expr = expr.replace(token, f"{col}{row}")

    if not required:
        return "=" + expr, []

    # Guarda: se qualquer celula exigida estiver vazia, NA() em vez de 0.
    guard = ",".join(f'{_DIMENSION_TOKENS[t][0]}{row}=""' for t in required)
    if len(required) > 1:
        guard = f"OR({guard})"
    return f"=IF({guard},NA(),{expr})", required


# ---------------------------------------------------------------------------
# 1) Geracao da aba de peca
# ---------------------------------------------------------------------------

def create_part_sheet(wb, template_sheet_name, new_sheet_name, part,
                       fastener_rows=3, tooling_rows=2):
    """
    Cria uma aba de peca nova, formatada como o template, com os blocos de
    Material e Processo preenchidos dinamicamente (quantas linhas forem
    necessarias) e os blocos de Fixador/Ferramental deixados em branco para
    preenchimento manual.

    `part` e um dicionario com:
        university, system, assembly, part_name, pn_base, suffix, details,
        qty,
        materials: lista de dicts com material, use, size1, unit1, size2,
                   unit2, area_name, area, length, density, quantity
        processes: lista de dicts com process, use, quantity

    Retorna um dict com o nome da aba e as celulas de Sub Total de cada
    bloco (materiais/processos/fixadores/ferramental), que a funcao de
    atualizacao da BOM mestra vai usar.
    """
    ws_template = wb[template_sheet_name]

    if new_sheet_name in wb.sheetnames:
        raise ValueError(f"Ja existe uma aba chamada '{new_sheet_name}'")

    ws = wb.copy_worksheet(ws_template)
    ws.title = new_sheet_name

    materials_catalog = _load_materials_catalog(wb)
    process_names = _load_catalog_names(wb, "Processes", 0)
    warnings = []

    # Copia largura de colunas do template (copy_worksheet ja traz isso,
    # mas garantimos aqui por seguranca)
    for col_letter, dim in ws_template.column_dimensions.items():
        ws.column_dimensions[col_letter].width = dim.width

    # Limpa tudo a partir da linha 9 (o cabecalho de peca, linhas 1-8, fica
    # igual ao template e sera sobrescrito com os dados da peca)
    if ws.max_row >= 9:
        ws.delete_rows(9, ws.max_row - 8)

    # --- Cabecalho (linhas 1-7) ---------------------------------------
    ws["B1"] = part.get("university", ws_template["B1"].value)
    ws["B2"] = part["system"]
    ws["B3"] = part["assembly"]
    ws["B4"] = part["part_name"]
    ws["B5"] = part["pn_base"]
    ws["B6"] = part["suffix"]
    ws["B7"] = part["details"]
    ws["N2"] = part["qty"]
    # K1 ('=Car'), K3 ('Go to BOM') sao mantidos como vieram do template
    # (referencia a um Named Range do workbook, nao mudam por peca)

    row = 9

    # --- Bloco de Materiais ---------------------------------------------
    mat_header_row = row
    _copy_row_style(ws_template, 9, ws, mat_header_row)
    headers_mat = ["ItemOrder", "Material", "Use", "UnitCost", "Size1",
                   "Unit1", "Size2", "Unit2", "Area Name", "Area", "Length",
                   "Density", "Quantity", "Sub Total"]
    for i, h in enumerate(headers_mat):
        ws.cell(row=mat_header_row, column=1 + i, value=h)
    row += 1

    mat_first_data_row = row
    materials = part.get("materials", [])
    for idx, m in enumerate(materials):
        r = mat_first_data_row + idx
        _copy_row_style(ws_template, 10, ws, r)
        ws.cell(row=r, column=1, value=(idx + 1) * 10)           # ItemOrder
        mat_name = m.get("material", "")
        ws.cell(row=r, column=2, value=mat_name)                 # Material
        ws.cell(row=r, column=3, value=m.get("use", ""))         # Use
        if mat_name and mat_name not in materials_catalog:
            warnings.append(
                f"Material '{mat_name}' (linha {r}, aba {new_sheet_name}) "
                f"nao encontrado no catalogo 'Materials' -- revisar nome."
            )
        uc_formula, required_tokens = _material_unitcost_formula(
            materials_catalog, mat_name, r)
        if uc_formula:
            ws.cell(row=r, column=4, value=uc_formula)
        # Confere agora, na geracao, se a peca forneceu tudo o que a formula
        # do catalogo exige -- ver _material_unitcost_formula. A guarda NA()
        # ja torna a falta visivel na planilha; este aviso a torna visivel
        # tambem para quem roda o pipeline, com nome, aba e linha, sem
        # precisar abrir o Excel para descobrir o que faltou.
        missing = [tok for tok in required_tokens
                   if m.get(_DIMENSION_TOKENS[tok][1]) is None]
        if missing:
            campos = ", ".join(_DIMENSION_TOKENS[t][1] for t in missing)
            warnings.append(
                f"Material '{mat_name}' (linha {r}, aba {new_sheet_name}): a "
                f"formula do catalogo exige {campos}, mas a peca nao forneceu "
                f"esse(s) valor(es). A celula de UnitCost vai mostrar #N/A ate "
                f"alguem preencher -- SEM esse preenchimento o custo deste "
                f"material nao entra no total."
            )
        ws.cell(row=r, column=5, value=m.get("size1"))
        ws.cell(row=r, column=6, value=m.get("unit1", ""))
        ws.cell(row=r, column=7, value=m.get("size2"))
        ws.cell(row=r, column=8, value=m.get("unit2", ""))
        ws.cell(row=r, column=9, value=m.get("area_name", ""))
        ws.cell(row=r, column=10, value=m.get("area"))
        ws.cell(row=r, column=11, value=m.get("length"))
        ws.cell(row=r, column=12, value=m.get("density"))
        ws.cell(row=r, column=13, value=m.get("quantity", 1))
        ws.cell(row=r, column=14,
                value=f'=IF(D{r}="","",D{r}*M{r})')

    n_mat = max(len(materials), 1)
    mat_subtotal_row = mat_first_data_row + n_mat
    _copy_row_style(ws_template, 14, ws, mat_subtotal_row)
    ws.cell(row=mat_subtotal_row, column=13, value="Sub Total")
    ws.cell(row=mat_subtotal_row, column=14,
            value=f"=SUM(N{mat_first_data_row}:N{mat_subtotal_row - 1})")

    row = mat_subtotal_row + 2  # 1 linha em branco de separacao

    # --- Bloco de Processos ----------------------------------------------
    proc_header_row = row
    _copy_row_style(ws_template, 16, ws, proc_header_row)
    headers_proc = ["ItemOrder", "Process", "Use", "UnitCost", "Unit",
                     "Quantity", "Multiplier Type", "Multiplier",
                     "Mult. Val.", "Sub Total"]
    for i, h in enumerate(headers_proc):
        ws.cell(row=proc_header_row, column=1 + i, value=h)
    row += 1

    proc_first_data_row = row
    processes = part.get("processes", [])
    mat_range = f"$A${mat_first_data_row}:$E${mat_subtotal_row - 1}"
    for idx, p in enumerate(processes):
        r = proc_first_data_row + idx
        _copy_row_style(ws_template, 17, ws, r)
        proc_name = p.get("process", "")
        ws.cell(row=r, column=1, value=(idx + 1) * 10)
        ws.cell(row=r, column=2, value=proc_name)
        ws.cell(row=r, column=3, value=p.get("use", ""))
        if proc_name and proc_name not in process_names:
            warnings.append(
                f"Processo '{proc_name}' (linha {r}, aba {new_sheet_name}) "
                f"nao encontrado no catalogo 'Processes' -- revisar nome."
            )
        ws.cell(row=r, column=4,
                value=f'=VLOOKUP(B{r},Processes!A:B,2,FALSE)')
        ws.cell(row=r, column=5,
                value=f'=IFERROR(VLOOKUP(B{r},Processes!A:C,3,FALSE),"")')
        ws.cell(row=r, column=6, value=p.get("quantity", 1))
        ws.cell(row=r, column=7,
                value=f'=VLOOKUP(B{r},Processes!A:F,6,FALSE)')
        mult_formula = (
            f'=IFERROR(IF(G{r}="Assembly",\'Process Multipliers\'!$C$3,'
            f'IF(G{r}="Fastener Installation",\'Process Multipliers\'!$C$5,'
            f'IF(G{r}="Drilling/Tapping",\'Process Multipliers\'!$C$7,'
            f'IF(G{r}="Material Machining, Forming",'
            f'IF(COUNTIF({mat_range},"*Steel*")>0,\'Process Multipliers\'!$C$21,'
            f'IF(COUNTIF({mat_range},"*Aluminum*")>0,\'Process Multipliers\'!$C$10,"")),'
            f'"")))),"")'
        )
        ws.cell(row=r, column=8, value=mult_formula)
        ws.cell(row=r, column=9,
                value=f"=IFERROR(VLOOKUP(H{r},'Process Multipliers'!C:E,3,FALSE),\"\")")
        ws.cell(row=r, column=10,
                value=f'=IF(I{r}<>"",D{r}*F{r}*I{r},D{r}*F{r})')

    n_proc = max(len(processes), 1)
    proc_subtotal_row = proc_first_data_row + n_proc
    _copy_row_style(ws_template, 25, ws, proc_subtotal_row)
    ws.cell(row=proc_subtotal_row, column=9, value="Sub Total")
    ws.cell(row=proc_subtotal_row, column=10,
            value=f"=SUM(J{proc_first_data_row}:J{proc_subtotal_row - 1})")

    row = proc_subtotal_row + 2

    # --- Bloco de Fixadores (preenchimento manual) ------------------------
    fast_header_row = row
    _copy_row_style(ws_template, 27, ws, fast_header_row)
    headers_fast = ["ItemOrder", "Fastener", "Use", "UnitCost", "Size1",
                     "Unit1", "Size2", "Unit2", "Quantity", "Sub Total"]
    for i, h in enumerate(headers_fast):
        ws.cell(row=fast_header_row, column=1 + i, value=h)
    row += 1
    fast_first_data_row = row
    for idx in range(fastener_rows):
        r = fast_first_data_row + idx
        _copy_row_style(ws_template, 28, ws, r)
        ws.cell(row=r, column=1, value=(idx + 1) * 10)
        ws.cell(row=r, column=10, value=f'=IF(D{r}="","",D{r}*I{r})')
    fast_subtotal_row = fast_first_data_row + fastener_rows
    _copy_row_style(ws_template, 29, ws, fast_subtotal_row)
    ws.cell(row=fast_subtotal_row, column=9, value="Sub Total")
    ws.cell(row=fast_subtotal_row, column=10,
            value=f"=SUM(J{fast_first_data_row}:J{fast_subtotal_row - 1})")

    row = fast_subtotal_row + 2

    # --- Bloco de Ferramental (preenchimento manual) -----------------------
    tool_header_row = row
    _copy_row_style(ws_template, 31, ws, tool_header_row)
    headers_tool = ["ItemOrder", "Tooling", "Use", "UnitCost", "Unit",
                     "Quantity", "PVF", "FracIncld", "Sub Total"]
    for i, h in enumerate(headers_tool):
        ws.cell(row=tool_header_row, column=1 + i, value=h)
    row += 1
    tool_first_data_row = row
    for idx in range(tooling_rows):
        r = tool_first_data_row + idx
        _copy_row_style(ws_template, 32, ws, r)
        ws.cell(row=r, column=1, value=(idx + 1) * 10)
        ws.cell(row=r, column=9,
                value=f'=IF(G{r}<>"",(D{r}*F{r}/G{r})*H{r},"")')
    tool_subtotal_row = tool_first_data_row + tooling_rows
    _copy_row_style(ws_template, 33, ws, tool_subtotal_row)
    ws.cell(row=tool_subtotal_row, column=8, value="Sub Total")
    ws.cell(row=tool_subtotal_row, column=9,
            value=f"=SUM(I{tool_first_data_row}:I{tool_subtotal_row - 1})")

    # --- Totais no cabecalho (N1 = Part Cost, N4 = Extended Cost) ---------
    ws["N1"] = (f"=N{mat_subtotal_row}+J{proc_subtotal_row}"
                f"+J{fast_subtotal_row}+I{tool_subtotal_row}")
    ws["N4"] = "=N1*N2"

    return {
        "sheet_name": new_sheet_name,
        "material_subtotal_cell": f"N{mat_subtotal_row}",
        "process_subtotal_cell": f"J{proc_subtotal_row}",
        "fastener_subtotal_cell": f"J{fast_subtotal_row}",
        "tooling_subtotal_cell": f"I{tool_subtotal_row}",
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# 2) Registro na aba BOM mestra + correcao de totais
# ---------------------------------------------------------------------------

def _scan_bom_blocks(ws):
    """
    Percorre a aba BOM e identifica os blocos por sistema:
    retorna lista de dicts {system, start_row, end_row, area_total_row}
    na ordem em que aparecem, alem do numero da linha 'Vehicle Total'.
    """
    blocks = []
    current_system = None
    start_row = None
    last_row = None
    vehicle_total_row = None

    for r in range(7, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value   # Area of Commodity
        f_val = ws.cell(row=r, column=6).value   # Component / marcador

        if b_val == "Vehicle Total":
            vehicle_total_row = r
            break

        if f_val == "Area Total":
            blocks.append({
                "system": current_system,
                "start_row": start_row,
                "end_row": last_row,
                "area_total_row": r,
            })
            current_system = None
            start_row = None
            last_row = None
            continue

        if b_val:
            if current_system is None:
                current_system = b_val
                start_row = r
            last_row = r

    return blocks, vehicle_total_row


def _renumber_line_nums(ws, first_row=7, last_row=None):
    """
    Reatribui o "Line Num." (coluna A) sequencialmente (1, 2, 3, ...) para
    cada linha de peca/sub-montagem, na ordem fisica em que aparecem na
    aba. Substitui qualquer numero antigo -- necessario porque inserir uma
    peca no meio da BOM desloca todas as linhas abaixo dela, e a numeracao
    precisa continuar refletindo a ordem de leitura (topo->baixo), nao a
    ordem/posicao em que cada peca foi inserida. Roda depois de toda
    insercao, junto com o recalculo dos Sub Totals/Area Totals.
    """
    if last_row is None:
        last_row = ws.max_row
    n = 0
    for r in range(first_row, last_row + 1):
        b_val = ws.cell(row=r, column=2).value
        if b_val == "Vehicle Total":
            break
        if ws.cell(row=r, column=6).value == "Area Total":
            continue
        if b_val:
            n += 1
            ws.cell(row=r, column=1, value=n)


def add_part_to_bom(wb, part, subtotals):
    """
    Insere uma nova linha na aba BOM referenciando a aba de peca gerada por
    create_part_sheet, dentro do bloco do sistema `part["system"]`, e depois
    reconstrói (do zero, a partir da estrutura real do arquivo pos-insercao)
    os Sub Totals de cada sistema, o Total do Veiculo e as referencias da
    aba Cost Summary -- evitando o problema de referencias de celula
    hardcoded ficarem desatualizadas apos o deslocamento de linhas.
    """
    ws_bom = wb["BOM"]
    target_system = part["system"]

    blocks, vehicle_total_row = _scan_bom_blocks(ws_bom)
    target_block = next((bl for bl in blocks if bl["system"] == target_system), None)
    if target_block is None:
        raise ValueError(
            f"Sistema '{target_system}' nao encontrado na aba BOM. "
            f"Sistemas disponiveis: {[b['system'] for b in blocks]}"
        )
    if target_block["start_row"] is None:
        raise ValueError(
            f"Sistema '{target_system}' existe na aba BOM mas ainda nao tem "
            f"nenhuma peca (bloco vazio) -- nao ha linha existente para "
            f"copiar estilo/formato, entao este script ainda nao sabe "
            f"inserir a PRIMEIRA peca de um sistema vazio. Adicionem essa "
            f"primeira peca manualmente uma vez; as proximas pecas desse "
            f"sistema poderao ser adicionadas automaticamente."
        )

    insert_at = target_block["area_total_row"]
    ws_bom.insert_rows(insert_at, 1)

    # Estiliza a nova linha copiando o estilo da ultima linha de peca daquele
    # bloco (garante fonte/cor/borda do sistema, como recomendado pelo Guide)
    _copy_row_style(ws_bom, target_block["end_row"], ws_bom, insert_at, max_col=15)

    sheet_name = subtotals["sheet_name"]

    ws_bom.cell(row=insert_at, column=2, value=part["system"])
    ws_bom.cell(row=insert_at, column=3, value=sheet_name)
    ws_bom.cell(row=insert_at, column=4, value=part["suffix"])
    ws_bom.cell(row=insert_at, column=6,
                value=f'=INDIRECT(C{insert_at}&"!B4")')
    ws_bom.cell(row=insert_at, column=7,
                value=f'=INDIRECT(C{insert_at}&"!B7")')
    ws_bom.cell(row=insert_at, column=8,
                value=f"=SUM(J{insert_at}:M{insert_at})")
    ws_bom.cell(row=insert_at, column=9,
                value=f"='{sheet_name}'!$N$2")
    ws_bom.cell(row=insert_at, column=10,
                value=f"='{sheet_name}'!{subtotals['material_subtotal_cell']}")
    ws_bom.cell(row=insert_at, column=11,
                value=f"='{sheet_name}'!{subtotals['process_subtotal_cell']}")
    ws_bom.cell(row=insert_at, column=12,
                value=f"='{sheet_name}'!{subtotals['fastener_subtotal_cell']}")
    ws_bom.cell(row=insert_at, column=13,
                value=f"='{sheet_name}'!{subtotals['tooling_subtotal_cell']}")
    ws_bom.cell(row=insert_at, column=14,
                value=f"=H{insert_at}*I{insert_at}")
    ws_bom.cell(row=insert_at, column=15, value=sheet_name)

    _fix_bom_and_cost_summary_totals(wb)


def _repair_self_referencing_formulas(ws, first_row=7, last_row=None):
    """
    Corrige formulas que auto-referenciam o proprio numero de linha em
    texto e ficaram desatualizadas apos um insert_rows -- ver os tres
    padroes acima (_INDIRECT_SELF_REF_PATTERN, _UNITCOST_SELF_REF_PATTERN,
    _TOTALCOST_SELF_REF_PATTERN). Reescreve como formula normal (nao-array)
    com o numero de linha correto -- no caso INDIRECT isso tambem elimina
    de quebra a fragilidade de metadado de "array formula" (ref
    desatualizado) que o openpyxl nao recalcula sozinho.
    """
    if last_row is None:
        last_row = ws.max_row
    fixed = 0
    for r in range(first_row, last_row + 1):
        for col in range(1, 16):  # colunas A..O
            cell = ws.cell(row=r, column=col)
            v = cell.value
            text = v.text if hasattr(v, "text") else (v if isinstance(v, str) and v.startswith("=") else None)
            if text is None:
                continue

            m = _INDIRECT_SELF_REF_PATTERN.search(text)
            if m:
                if int(m.group(1)) != r:
                    cell.value = _INDIRECT_SELF_REF_PATTERN.sub(f"C{r}&", text)
                    fixed += 1
                continue

            m = _UNITCOST_SELF_REF_PATTERN.fullmatch(text)
            if m:
                if int(m.group(1)) != r or int(m.group(2)) != r:
                    cell.value = f"=SUM(J{r}:M{r})"
                    fixed += 1
                continue

            m = _TOTALCOST_SELF_REF_PATTERN.fullmatch(text)
            if m:
                if int(m.group(1)) != r or int(m.group(2)) != r:
                    cell.value = f"=H{r}*I{r}"
                    fixed += 1
                continue
    return fixed


def _fix_bom_and_cost_summary_totals(wb):
    ws_bom = wb["BOM"]
    _repair_self_referencing_formulas(ws_bom)
    _renumber_line_nums(ws_bom)
    blocks, vehicle_total_row = _scan_bom_blocks(ws_bom)

    for bl in blocks:
        r = bl["area_total_row"]
        s, e = bl["start_row"], bl["end_row"]
        for col_letter in ["J", "K", "L", "M"]:
            ws_bom[f"{col_letter}{r}"] = (
                f"=SUMPRODUCT($I{s}:$I{e},{col_letter}{s}:{col_letter}{e})"
            )
        ws_bom[f"N{r}"] = f"=SUM(J{r}:M{r})"

    if vehicle_total_row is not None:
        for col_letter in ["J", "K", "L", "M", "N"]:
            terms = "+".join(f"{col_letter}{bl['area_total_row']}" for bl in blocks)
            ws_bom[f"{col_letter}{vehicle_total_row}"] = f"={terms}"
        ws_bom["O1"] = f"=N{vehicle_total_row}"

    # Aba Cost Summary: linhas 7 em diante seguem a MESMA ordem dos sistemas
    # da BOM (confirmado comparando os dois arquivos)
    ws_cs = wb["Cost Summary"]
    cs_row = 7
    for bl in blocks:
        r = bl["area_total_row"]
        for col_letter, bom_col in zip(["D", "E", "F", "G", "H"],
                                        ["J", "K", "L", "M", "N"]):
            ws_cs[f"{col_letter}{cs_row}"] = f"=BOM!{bom_col}{r}"
        cs_row += 1
